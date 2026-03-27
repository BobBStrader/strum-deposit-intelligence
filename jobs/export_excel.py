"""
Excel Export for Deposit Ranking Report
Usage:
    python3 export_excel.py --client "Securityplus FCU" --market "Baltimore" MD --output /tmp/report.xlsx
"""

import argparse
import os
import sys
from datetime import date

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from deposit_ranking_report import (
    get_market_rates, get_market_liquid_rates, find_client,
    build_liquid_table, build_term_table,
    CD_TERMS, CD_LABELS, MIN_BALANCES, MIN_BAL_LABELS,
    LIQUID_PRODUCTS, LIQUID_LABELS
)
from scrapers.schema import get_conn
from scrapers.peer_group import get_peers

# ── Colors ────────────────────────────────────────────────────────────────────
NAVY        = "0D2B55"
NAVY2       = "1A3F6F"
GOLD        = "C8922A"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
RED_HEX     = "CC0000"
GREEN_HEX   = "006400"

def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def make_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def write_table(ws, rows, title, start_row):
    """Write a single ranked product table. Returns next available row."""
    # Section header
    tc = ws.cell(row=start_row, column=1, value=title)
    tc.font = Font(bold=True, size=11, color=WHITE)
    tc.fill = make_fill(NAVY)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    ws.row_dimensions[start_row].height = 18
    start_row += 1

    # Column headers
    for ci, h in enumerate(["Rank", "Institution", "APY %", "Chg (bps)"], 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = Font(bold=True, size=9, color=WHITE)
        c.fill = make_fill(NAVY2)
        c.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")
    start_row += 1

    # Assign ranks
    rank = 1
    for r in rows:
        r["_rank"] = rank if r["apy"] is not None else "N/O"
        if r["apy"] is not None:
            rank += 1

    # Data rows
    for i, row in enumerate(rows):
        is_client = row.get("is_client", False)
        fill = make_fill(GOLD) if is_client else (make_fill(LIGHT_GRAY) if i % 2 == 0 else make_fill(WHITE))

        rank_val = row["_rank"]
        name_val = ("► " if is_client else "") + row["name"]
        apy_val  = f"{row['apy']:.2f}%" if row["apy"] is not None else "N/O"
        chg_val  = ""
        if row["apy"] is not None and row.get("prior_apy") is not None:
            bps = round((row["apy"] - row["prior_apy"]) * 100, 1)
            chg_val = f"+{bps}" if bps > 0 else str(bps)

        for ci, val in enumerate([rank_val, name_val, apy_val, chg_val], 1):
            cell = ws.cell(row=start_row, column=ci, value=val)
            cell.fill = fill
            cell.font = Font(size=9, bold=is_client)
            cell.alignment = Alignment(horizontal="center" if ci in (1, 3, 4) else "left", vertical="center")
            cell.border = make_border()

        # Color the change cell
        if chg_val:
            chg_cell = ws.cell(row=start_row, column=4)
            try:
                v = float(chg_val.replace("+", ""))
                chg_cell.font = Font(size=9, bold=is_client, color=(GREEN_HEX if v > 0 else RED_HEX))
            except ValueError:
                pass

        start_row += 1

    # Blank spacer
    start_row += 1
    return start_row


def build_excel(client_name, city, state, output_path):
    conn = get_conn()
    peers = get_peers(conn, city, state)
    if not peers:
        print(f"No peers found for {city}, {state}")
        return

    client = find_client(peers, client_name)
    client_id = client["institution_id"] if client else None

    cd_rates, cd_meta   = get_market_rates(conn, city, state)
    liq_rates, liq_meta = get_market_liquid_rates(conn, city, state)

    today = date.today().strftime("%m/%d/%Y")
    wb = openpyxl.Workbook()

    # ── Cover sheet ───────────────────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.column_dimensions["A"].width = 50

    ws_cover.cell(row=2, column=1, value="Deposit Ranking Report").font = Font(bold=True, size=20, color=NAVY)
    ws_cover.cell(row=4, column=1, value=f"Client: {client_name}").font = Font(size=13)
    ws_cover.cell(row=5, column=1, value=f"Market: {city}, {state}").font = Font(size=13)
    ws_cover.cell(row=6, column=1, value=f"Generated: {today}").font = Font(size=11, color="666666")
    ws_cover.cell(row=7, column=1, value=f"Peers: {len(peers)} institutions").font = Font(size=11, color="666666")
    ws_cover.cell(row=9, column=1, value="Powered by Strum Platform").font = Font(size=10, italic=True, color="999999")

    # ── Liquid Rates sheet ────────────────────────────────────────────────────
    ws_liq = wb.create_sheet("Liquid Rates")
    ws_liq.column_dimensions["A"].width = 6
    ws_liq.column_dimensions["B"].width = 40
    ws_liq.column_dimensions["C"].width = 10
    ws_liq.column_dimensions["D"].width = 12

    row = 1
    for product in LIQUID_PRODUCTS:
        rows = build_liquid_table(liq_meta, liq_rates, product, client_id)
        row = write_table(ws_liq, rows, LIQUID_LABELS[product], row)

    # ── CD sheets (one per min balance) ───────────────────────────────────────
    for min_bal in MIN_BALANCES:
        ws_cd = wb.create_sheet(f"CDs {MIN_BAL_LABELS[min_bal]}")
        ws_cd.column_dimensions["A"].width = 6
        ws_cd.column_dimensions["B"].width = 40
        ws_cd.column_dimensions["C"].width = 10
        ws_cd.column_dimensions["D"].width = 12

        row = 1
        for term in CD_TERMS:
            rows = build_term_table(cd_meta, cd_rates, term, min_bal, client_id)
            if not rows:
                continue
            label = f"{CD_LABELS[term]} — {MIN_BAL_LABELS[min_bal]} min"
            row = write_table(ws_cd, rows, label, row)

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel Deposit Ranking Report")
    parser.add_argument("--client", required=True)
    parser.add_argument("--market", nargs=2, metavar=("CITY", "STATE"), required=True)
    parser.add_argument("--output", default="/tmp/deposit_ranking_report.xlsx")
    args = parser.parse_args()

    build_excel(args.client, args.market[0], args.market[1], args.output)
